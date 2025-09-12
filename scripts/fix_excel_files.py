#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修复Excel文件格式问题的脚本
"""

import os
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import Workbook

def fix_excel_file(input_path, output_path):
    """修复单个Excel文件"""
    try:
        print(f"🔧 正在修复文件: {input_path}")
        
        # 方法1: 尝试用openpyxl直接读取并重新保存
        try:
            wb = openpyxl.load_workbook(input_path, data_only=True)
            ws = wb.active
            
            # 提取数据
            data = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append(row)
            
            if not data:
                print(f"⚠️ 文件 {input_path} 没有数据")
                return False
            
            # 创建新的工作簿
            new_wb = Workbook()
            new_ws = new_wb.active
            
            # 写入数据
            for i, row in enumerate(data):
                for j, cell in enumerate(row):
                    new_ws.cell(row=i+1, column=j+1, value=cell)
            
            # 保存修复后的文件
            new_wb.save(output_path)
            print(f"✅ 成功修复并保存到: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ 修复失败: {str(e)}")
            return False
            
    except Exception as e:
        print(f"❌ 处理文件时出错: {str(e)}")
        return False

def convert_to_csv(input_path, output_path):
    """将Excel文件转换为CSV"""
    try:
        print(f"🔄 正在转换为CSV: {input_path}")
        
        # 使用openpyxl读取
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws = wb.active
        
        # 提取数据
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        
        if not data:
            print(f"⚠️ 文件 {input_path} 没有数据")
            return False
        
        # 创建DataFrame
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # 保存为CSV
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"✅ 成功转换为CSV: {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ 转换失败: {str(e)}")
        return False

def main():
    """主函数"""
    excel_dir = Path("scripts/publish_excel")
    fixed_dir = Path("scripts/publish_excel_fixed")
    csv_dir = Path("scripts/publish_excel_csv")
    
    # 创建输出目录
    fixed_dir.mkdir(exist_ok=True)
    csv_dir.mkdir(exist_ok=True)
    
    print("🚀 开始修复Excel文件...")
    print("=" * 50)
    
    # 获取所有Excel文件
    excel_files = list(excel_dir.glob("*.xlsx")) + list(excel_dir.glob("*.xls"))
    
    if not excel_files:
        print("⚠️ 未找到Excel文件")
        return
    
    print(f"📁 找到 {len(excel_files)} 个Excel文件")
    
    success_count = 0
    
    for excel_file in excel_files:
        print(f"\n📖 处理文件: {excel_file.name}")
        
        # 尝试修复Excel文件
        fixed_path = fixed_dir / excel_file.name
        if fix_excel_file(excel_file, fixed_path):
            success_count += 1
        
        # 同时转换为CSV
        csv_path = csv_dir / (excel_file.stem + ".csv")
        convert_to_csv(excel_file, csv_path)
    
    print("\n" + "=" * 50)
    print(f"✅ 处理完成！成功修复 {success_count} 个文件")
    print(f"📁 修复后的Excel文件保存在: {fixed_dir}")
    print(f"📁 CSV文件保存在: {csv_dir}")

if __name__ == "__main__":
    main()
